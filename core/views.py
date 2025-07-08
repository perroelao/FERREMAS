import requests
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from django.db import connection
from django.views.decorators.http import require_POST
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.auth import logout as django_logout
from .paypal_utils import *
from .models import *
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json


def home(request):
    return render(request, 'core/index.html')

def carrito(request):
    return render(request, 'core/cart.html')

def listaproductos(request):
    url = 'http://localhost:3002/productos/'
    response = requests.get(url)
    
    if response.status_code == 200:
        productos = response.json()
        return render(request, 'core/listaproductos.html', {'productos': productos})
    else:
        messages.error(request, 'Error al cargar los productos')
        return redirect('home')  # Redirigir a la página de inicio en caso de error
    
def aprobacion_pagos(request):
    # Solo pagos transferencia, estado pendiente, pedido pendiente
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT p.pago_id, p.pedido_id, p.monto, p.fecha_pago, p.estado_pago_id, ep.nombre AS estado_pago,
                   p.metodo_pago_id, mp.nombre AS metodo_pago, p.transaccion_id, p.detalle,
                   ped.cliente_id, u.nombre || ' ' || u.apellido_p AS cliente
            FROM PAGO p
            JOIN ESTADO_PAGO ep ON p.estado_pago_id = ep.estado_pago_id
            JOIN METODO_PAGO mp ON p.metodo_pago_id = mp.metodo_pago
            JOIN PEDIDO ped ON p.pedido_id = ped.pedido_id
            JOIN USUARIO u ON ped.cliente_id = u.id_usuario
            WHERE p.metodo_pago_id = 4
              AND p.estado_pago_id = 2
              AND ped.estado_id = 1
            ORDER BY p.fecha_pago DESC
        """)
        pagos = dictfetchall(cursor)
    return render(request, 'core/aprobacion_pagos.html', {'pagos': pagos})

def dictfetchall(cursor):
    "Devuelve todos los resultados como una lista de diccionarios"
    columns = [col[0].lower() for col in cursor.description]
    return [
        dict(zip(columns, row))
        for row in cursor.fetchall()
    ]

@require_POST
def aprobar_pago(request, pago_id):
    with connection.cursor() as cursor:
        cursor.execute("UPDATE PAGO SET estado_pago_id = 1 WHERE pago_id = %s", [pago_id])  # 3 = Aprobado
    messages.success(request, 'Pago aprobado correctamente')
    return redirect('aprobacion_pagos')

@require_POST
def rechazar_pago(request, pago_id):
    with connection.cursor() as cursor:
        cursor.execute("UPDATE PAGO SET estado_pago_id = 3 WHERE pago_id = %s", [pago_id])  # 4 = Rechazado
    messages.success(request, 'Pago rechazado correctamente')
    return redirect('aprobacion_pagos')

def gestion_pedidos(request):
    # Pedidos que no estén completados ni cancelados
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT ped.pedido_id, ped.fecha_pedido, ped.estado_id, e.nombre AS estado,
                   ped.cliente_id, u.nombre || ' ' || u.apellido_p AS cliente, ped.total
            FROM PEDIDO ped
            JOIN ESTADO e ON ped.estado_id = e.estado_id
            JOIN USUARIO u ON ped.cliente_id = u.id_usuario
            WHERE ped.estado_id = 1 -- Pendiente
            ORDER BY ped.fecha_pedido DESC
        """)
        pedidos = dictfetchall(cursor)
    return render(request, 'core/gestion_pedidos.html', {'pedidos': pedidos})

@require_POST
def aprobar_pedido(request, pedido_id):
    # Estado 3 = Aprobado (ajusta según tu tabla ESTADO)
    with connection.cursor() as cursor:
        cursor.execute("UPDATE PEDIDO SET estado_id = 3 WHERE pedido_id = %s", [pedido_id])
    messages.success(request, 'Pedido aprobado correctamente')
    return redirect('gestion_pedidos')

@require_POST
def rechazar_pedido(request, pedido_id):
    # Estado 4 = Rechazado/Cancelado (ajusta según tu tabla ESTADO)
    with connection.cursor() as cursor:
        cursor.execute("UPDATE PEDIDO SET estado_id = 4 WHERE pedido_id = %s", [pedido_id])
    messages.success(request, 'Pedido rechazado correctamente')
    return redirect('gestion_pedidos')

def admin_usuarios(request):
    # Obtener usuarios desde la API
    response = requests.get('http://localhost:8001/usuarios/')
    usuarios = []
    if response.status_code == 200:
        usuarios = response.json()
    return render(request, 'core/admin_usuarios.html', {'usuarios': usuarios})
    
def agregar_usuario(request):
    if request.method == 'POST':
        rol_id = int(request.POST.get('rol_id'))
        deber = 0  # Por defecto 0 si no se envía
        if rol_id == 3:  # Rol de administrador
            deber = 1  # Administradores deben cambiar la contraseña al iniciar sesión por primera vez
        data = {
            'rut': request.POST.get('rut', ''),
            'nombre': request.POST.get('nombre', ''),
            'apellido_p': request.POST.get('apellido_p', ''),
            'apellido_m': request.POST.get('apellido_m', ''),
            'snombre': request.POST.get('snombre', ''),
            'email': request.POST.get('email', ''),
            'fono': request.POST.get('fono', ''),
            'direccion': request.POST.get('direccion', ''),
            'password': request.POST.get('password', ''),
            'debe_cambiar_password': deber,  # 1 si debe cambiar la contraseña, 0 si no
            'rol_id': rol_id,
            
        }
        print("DATA ENVIADA:", data)
        response = requests.post('http://localhost:8001/usuarios/add_user/', params=data)
        print("RESPUESTA:", response.status_code, response.text)
        if response.status_code == 201 or response.status_code == 200:
            messages.success(request, 'Usuario agregado correctamente')
        else:
            messages.error(request, f'Error al agregar usuario: {response.text}')
        return redirect('admin_usuarios')

    response = requests.get('http://localhost:8001/usuarios/')
    usuarios = []
    if response.status_code == 200:
        usuarios = response.json()
    return render(request, 'core/admin_usuarios.html', {'usuarios': usuarios})
   

def editar_usuario(request, id_actualizar):
    if request.method == 'POST':
        data = {
            'rut': request.POST['rut'],
            'nombre': request.POST['nombre'],
            'apellido_p': request.POST['apellido_p'],
            'apellido_m': request.POST['apellido_m'],
            'snombre': request.POST['snombre'],
            'email': request.POST['email'],
            'fono': request.POST['telefono'],
            'direccion': request.POST['direccion'],
            'password': request.POST['password'],
            'rol_id': int(request.POST['rol']),
        }
        url = f'http://localhost:8001/usuarios/{id_actualizar}'
        response = requests.put(url, params=data)
        if response.status_code == 200:
            messages.success(request, 'Usuario actualizado correctamente')
        else:
            messages.error(request, 'Error al actualizar usuario')
    return redirect('admin_usuarios')

def eliminar_usuario(request, id_eliminar):
    url = f'http://localhost:8001/usuarios/{id_eliminar}'
    response = requests.delete(url)
    if response.status_code == 200:
        messages.success(request, 'Usuario eliminado correctamente')
    else:
        messages.error(request, 'Error al eliminar usuario')
    return redirect('admin_usuarios')

def login(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        try:
            response = requests.post(
                'http://localhost:8001/usuarios/login/',
                params={'email': email, 'password': password}
            )
            if response.status_code == 200:
                data = response.json()
                request.session['usuario_id'] = data.get('id_usuario')
                request.session['nombre'] = data.get('nombre')
                request.session['rol_id'] = data.get('rol_id')
                rol = data.get('rol_id')
                # Redirección según rol
                if rol == 1 or rol == 2:
                    return redirect('home')
                elif rol == 3:
                    if data.get('rol_id') == 3 and data.get('debe_cambiar_password', 0) == 1:
                        request.session['usuario_id'] = data.get('id_usuario')
                        return redirect('cambiar_password_admin')
                    return redirect('admin_usuarios')
                elif rol == 4:
                    return redirect('gestion_pedidos')
                elif rol == 5:
                    return redirect('aprobacion_pagos')
                else:
                    messages.error(request, 'Rol no reconocido.')
            else:
                messages.error(request, 'Correo o contraseña incorrectos.')
        except Exception as e:
            messages.error(request, f'Error de conexión: {e}')
    return render(request, 'core/login.html')

def logout(request):
    django_logout(request)  # Limpia la sesión
    return redirect('home')

def detalle_producto(request, producto_id):
    try:
        resp = requests.get(f'http://localhost:3002/productos/{producto_id}')
        resp.raise_for_status()
        producto = resp.json()
    except Exception:
        producto = None
    return render(request, 'core/detalle_producto.html', {'producto': producto})

def confirmar_pedido(request):
    return render(request, 'core/confirmar_pedido.html')

def dictfetchall(cursor):
    columns = [col[0].lower() for col in cursor.description]
    return [
        dict(zip(columns, row))
        for row in cursor.fetchall()
    ]

def informe_pedidos(request):
    # Solo permitir acceso a administradores (rol_id == 3)
    if request.session.get('rol_id') != 3:
        return redirect('home')
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT ped.pedido_id, ped.fecha_pedido, ped.total, e.nombre AS estado, u.nombre || ' ' || u.apellido_p AS cliente
            FROM PEDIDO ped
            JOIN ESTADO e ON ped.estado_id = e.estado_id
            JOIN USUARIO u ON ped.cliente_id = u.id_usuario
            ORDER BY ped.fecha_pedido DESC
        """)
        pedidos = dictfetchall(cursor)
    return render(request, 'core/informe_pedidos.html', {'pedidos': pedidos})

def descargar_informe_pedidos(request):
    if request.session.get('rol_id') != 3:
        return redirect('home')
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT ped.pedido_id, ped.fecha_pedido, ped.total, e.nombre AS estado, u.nombre || ' ' || u.apellido_p AS cliente
            FROM PEDIDO ped
            JOIN ESTADO e ON ped.estado_id = e.estado_id
            JOIN USUARIO u ON ped.cliente_id = u.id_usuario
            ORDER BY ped.fecha_pedido DESC
        """)
        pedidos = dictfetchall(cursor)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos"
    headers = ["ID Pedido", "Fecha", "Cliente", "Estado", "Total"]
    ws.append(headers)
    for p in pedidos:
        ws.append([
            p['pedido_id'],
            str(p['fecha_pedido']),
            p['cliente'],
            p['estado'],
            float(p['total'])
        ])
    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=pedidos.xlsx'
    wb.save(response)
    return response

def informe_pagos(request):
    # Solo permitir acceso a administradores (rol_id == 3)
    if request.session.get('rol_id') != 3:
        return redirect('home')
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT p.pago_id, p.pedido_id, p.monto, p.fecha_pago, ep.nombre AS estado_pago,
                   mp.nombre AS metodo_pago, u.nombre || ' ' || u.apellido_p AS cliente
            FROM PAGO p
            JOIN ESTADO_PAGO ep ON p.estado_pago_id = ep.estado_pago_id
            JOIN METODO_PAGO mp ON p.metodo_pago_id = mp.metodo_pago
            JOIN PEDIDO ped ON p.pedido_id = ped.pedido_id
            JOIN USUARIO u ON ped.cliente_id = u.id_usuario
            ORDER BY p.fecha_pago DESC
        """)
        pagos = dictfetchall(cursor)
    return render(request, 'core/informe_pagos.html', {'pagos': pagos})

def descargar_informe_pagos(request):
    if request.session.get('rol_id') != 3:
        return redirect('home')
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT p.pago_id, p.pedido_id, p.monto, p.fecha_pago, ep.nombre AS estado_pago,
                   mp.nombre AS metodo_pago, u.nombre || ' ' || u.apellido_p AS cliente
            FROM PAGO p
            JOIN ESTADO_PAGO ep ON p.estado_pago_id = ep.estado_pago_id
            JOIN METODO_PAGO mp ON p.metodo_pago_id = mp.metodo_pago
            JOIN PEDIDO ped ON p.pedido_id = ped.pedido_id
            JOIN USUARIO u ON ped.cliente_id = u.id_usuario
            ORDER BY p.fecha_pago DESC
        """)
        pagos = dictfetchall(cursor)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pagos"
    headers = ["ID Pago", "ID Pedido", "Cliente", "Método de Pago", "Estado Pago", "Monto", "Fecha"]
    ws.append(headers)
    for p in pagos:
        ws.append([
            p['pago_id'],
            p['pedido_id'],
            p['cliente'],
            p['metodo_pago'],
            p['estado_pago'],
            float(p['monto']),
            str(p['fecha_pago'])
        ])
    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=pagos.xlsx'
    wb.save(response)
    return response

def pedidos_para_entregar(request):
    # Solo mostrar si el usuario es vendedor
    if request.session.get('rol_id') != 2:
        return render(request, 'core/no_autorizado.html')

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT p.pedido_id, u.nombre || ' ' || u.apellido_p AS cliente, p.fecha_pedido, p.total, e.nombre AS estado
            FROM PEDIDO p
            JOIN USUARIO u ON p.cliente_id = u.id_usuario
            JOIN ESTADO e ON p.estado_id = e.estado_id
            JOIN PAGO pa ON pa.pedido_id = p.pedido_id
            WHERE p.estado_id = 3 -- Completado
              AND pa.estado_pago_id = 1 -- Pagado
            ORDER BY p.fecha_pedido DESC
        """)
        pedidos = [
            {
                'pedido_id': row[0],
                'cliente': row[1],
                'fecha_pedido': row[2],
                'total': row[3],
                'estado': row[4],
            }
            for row in cursor.fetchall()
        ]
    return render(request, 'core/pedidos_para_entregar.html', {'pedidos': pedidos})

@require_POST
def marcar_pedido_entregado(request, pedido_id):
    with connection.cursor() as cursor:
        cursor.execute("UPDATE PEDIDO SET estado_id = 2 WHERE pedido_id = %s", [pedido_id])  # 2 = Entregado
    messages.success(request, 'Pedido marcado como entregado correctamente')
    return redirect('pedidos_para_entregar')

def gestion_pedidos(request):
    # Pedidos que no estén completados ni cancelados
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT ped.pedido_id, ped.fecha_pedido, ped.estado_id, e.nombre AS estado,
                   ped.cliente_id, u.nombre || ' ' || u.apellido_p AS cliente, ped.total
            FROM PEDIDO ped
            JOIN ESTADO e ON ped.estado_id = e.estado_id
            JOIN USUARIO u ON ped.cliente_id = u.id_usuario
            WHERE ped.estado_id = 1
            ORDER BY ped.fecha_pedido DESC
        """)
        pedidos = dictfetchall(cursor)

    # --- NUEVO: Obtener productos desde la API ---
    productos = []
    try:
        r = requests.get("http://localhost:3002/productos")
        if r.status_code == 200:
            productos = r.json()
    except Exception as e:
        productos = []
    # --- FIN NUEVO ---

    # También necesitas pasar categorias y marcas si usas los selects
    categorias = Categoria.objects.all()
    marcas = Marca.objects.all()

    return render(request, 'core/gestion_pedidos.html', {
        'pedidos': pedidos,
        'productos': productos,
        'categorias': categorias,
        'marcas': marcas,
    })

def get_next_pedido_id():
    with connection.cursor() as cursor:
        cursor.execute("SELECT PEDIDO_SEQ.NEXTVAL FROM DUAL")
        row = cursor.fetchone()
    return row[0]

def get_next_detalle_id():
    with connection.cursor() as cursor:
        cursor.execute("SELECT DETALLE_PEDIDO_SEQ.NEXTVAL FROM DUAL")
        row = cursor.fetchone()
    return row[0]

def get_next_pago_id():
    from django.db import connection
    with connection.cursor() as cursor:
        cursor.execute("SELECT PAGO_SEQ.NEXTVAL FROM DUAL")
        row = cursor.fetchone()
    return row[0]


@csrf_exempt
def crear_pedido_paypal(request):
    if request.method == "POST":
        data = json.loads(request.body)
        total = data.get("total")
        cliente_id = data.get("cliente_id")
        sucursal_retiro = data.get("sucursal_retiro")
        vendedor_id = data.get("vendedor_id")
        productos = data.get("productos")
        metodo_pago_id = data.get("metodo_pago_id")
        tipo_despacho = data.get("tipo_despacho")  # "domicilio" o "retiro"

        # Validación básica
        if not total or not cliente_id or not productos or not metodo_pago_id or not tipo_despacho:
            return JsonResponse({"error": "Datos incompletos"}, status=400)

        # Mapear tipo_despacho a su ID
        if tipo_despacho == "domicilio":
            tipo_despacho_id = 1
            sucursal = None
        elif tipo_despacho == "retiro":
            tipo_despacho_id = 2
            if not sucursal_retiro:
                return JsonResponse({"error": "Debes seleccionar una sucursal para retiro"}, status=400)
            try:
                sucursal = Sucursal.objects.get(sucursal_id=sucursal_retiro)
            except Sucursal.DoesNotExist:
                return JsonResponse({"error": "Sucursal no encontrada"}, status=400)
        else:
            return JsonResponse({"error": "Tipo de despacho inválido"}, status=400)

        try:
            cliente = Usuario.objects.get(id_usuario=cliente_id)

            # Buscar vendedor: primero rol_id=2, si no hay, usa admin (rol_id=3)
            vendedor = None
            if vendedor_id:
                try:
                    vendedor = Usuario.objects.get(id_usuario=vendedor_id)
                except Usuario.DoesNotExist:
                    vendedor = None
            if not vendedor:
                vendedor = Usuario.objects.filter(rol_id=2).first()
            if not vendedor:
                vendedor = Usuario.objects.filter(rol_id=3).first()
            if not vendedor:
                return JsonResponse({"error": "No hay vendedor ni administrador disponible para asignar al pedido."}, status=400)

            # ESTADOS FIJOS
            estado_pedido = Estado.objects.get(estado_id=1)  # Pendiente
            if metodo_pago_id == 1:  # PayPal
                estado_pago = EstadoPago.objects.get(estado_pago_id=1)  # Pagado
            else:  # Transferencia
                estado_pago = EstadoPago.objects.get(estado_pago_id=2)  # Pendiente

            metodo_pago = MetodoPago.objects.get(metodo_pago=metodo_pago_id)
            tipo_despacho_obj = TipoDespacho.objects.get(pk=tipo_despacho_id)

            # 1. Crea el pedido
            nuevo_id = get_next_pedido_id()
            pedido = Pedido.objects.create(
                pedido_id=nuevo_id,
                cliente=cliente,
                sucursal=sucursal if tipo_despacho_id == 2 else None,
                vendedor=vendedor,
                estado=estado_pedido,
                total=total,
                tipo_despacho=tipo_despacho_obj,
            )

            # 2. Crea los detalles del pedido y descuenta stock
            for prod in productos:
                producto = Producto.objects.get(producto_id=prod["producto_id"])
                if producto.stock < prod["cantidad"]:
                    return JsonResponse({"error": f"Stock insuficiente para {producto.nombre}"}, status=400)
                detalle_id = get_next_detalle_id()
                DetallePedido.objects.create(
                    detalle_id=detalle_id,
                    pedido=pedido,
                    producto=producto,
                    cantidad=prod["cantidad"],
                    precio_unit=prod["precio_unit"]
                )
                producto.stock -= prod["cantidad"]
                producto.save()

            # 3. Crea el pago según método
            pago_id = get_next_pago_id()
            if metodo_pago_id == 1:
                # PayPal: crea orden y asocia transaccion_id
                resultado = crear_orden_paypal(total)
                Pago.objects.create(
                    pago_id=pago_id,
                    pedido=pedido,
                    metodo_pago=metodo_pago,
                    monto=total,
                    estado_pago=estado_pago,
                    transaccion_id=resultado["orderId"]
                )
                resultado["pedido_id"] = pedido.pedido_id
                return JsonResponse(resultado)
            elif metodo_pago_id == 4:
                comprobante = data.get("comprobante")  # base64 string o None
                Pago.objects.create(
                    pago_id=pago_id,
                    pedido=pedido,
                    metodo_pago=metodo_pago,
                    monto=total,
                    estado_pago=estado_pago,
                    detalle=comprobante
                )
                return JsonResponse({"pedido_id": pedido.pedido_id})
            else:
                return JsonResponse({"error": "Método de pago no soportado"}, status=400)

        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({"error": str(e)}, status=500)
    return JsonResponse({"error": "Método no permitido"}, status=405)

@csrf_exempt
def capturar_pago_paypal_view(request):
    if request.method == "POST":
        data = json.loads(request.body)
        order_id = data.get("orderId")
        if not order_id:
            return JsonResponse({"error": "Falta orderId"}, status=400)
        try:
            # 1. Captura el pago en PayPal
            resultado = capturar_pago_paypal(order_id)

            # 2. Busca el pago y pedido en tu base de datos
            pago = Pago.objects.get(transaccion_id=order_id)
            pedido = pago.pedido

            # 3. Actualiza el estado del pago y pedido
            estado_pago_pagado = EstadoPago.objects.get(nombre__iexact="pagado")
            #estado_pedido_completado = Estado.objects.get(nombre__iexact="completado")
            pago.estado_pago = estado_pago_pagado
            pago.save()
            #pedido.estado = estado_pedido_completado
            pedido.save()

            return JsonResponse({"status": "Pago capturado y pedido actualizado", "paypal": resultado})
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({"error": str(e)}, status=500)
    return JsonResponse({"error": "Método no permitido"}, status=405)

def pago_exitoso(request):
    return render(request, 'core/pago_exitoso.html')

def detalle_pedido_ajax(request, pedido_id):
    pedido = Pedido.objects.get(pedido_id=pedido_id)
    detalles = DetallePedido.objects.filter(pedido=pedido)
    productos = [
        {
            "nombre": d.producto.nombre,
            "cantidad": d.cantidad,
            "precio_unit": d.precio_unit
        }
        for d in detalles
    ]
    return JsonResponse({
        "productos": productos
    })

def detalle_pago_ajax(request, pago_id):
    pago = Pago.objects.get(pago_id=pago_id)
    return JsonResponse({
        "foto": pago.detalle  # Suponiendo que aquí guardas la URL o path de la foto
    })

def detalle_entrega_ajax(request, pedido_id):
    pedido = Pedido.objects.get(pedido_id=pedido_id)
    return JsonResponse({
        "cliente_id": pedido.cliente.id_usuario,
        "cliente_nombre": f"{pedido.cliente.nombre} {pedido.cliente.apellido_p}",
        "pedido_id": pedido.pedido_id,
        "direccion_entrega": pedido.cliente.direccion,  # o el campo que uses
        "tipo_despacho": pedido.tipo_despacho.nombre,
        "sucursal": pedido.sucursal.nombre if pedido.sucursal else None,

    })

def agregar_producto(request):
    if request.method == "POST":
        nombre = request.POST.get("nombre")
        descripcion = request.POST.get("descripcion")
        precio = request.POST.get("precio")
        stock = request.POST.get("stock")
        categoria_id = request.POST.get("categoria_id")
        marca_id = request.POST.get("marca_id")

        # Procesar imagen subida por el usuario
        imagen_file = request.FILES.get("imagen_file")
        imagen_base64 = ""
        if imagen_file:
            imagen_base64 = "data:" + imagen_file.content_type + ";base64," + base64.b64encode(imagen_file.read()).decode()
        else:
            imagen_base64 = request.POST.get("imagen", "")

        payload = {
            "nombre": nombre,
            "descripcion": descripcion,
            "precio": int(precio),
            "stock": int(stock),
            "imagen": imagen_base64,
            "categoria_id": int(categoria_id),
            "marca_id": int(marca_id)
        }

        try:
            r = requests.post("http://localhost:3002/crear_productos", json=payload)
            if r.status_code == 201:
                messages.success(request, "Producto agregado correctamente.")
            else:
                messages.error(request, f"Error al agregar producto: {r.json().get('error', 'Error desconocido')}")
        except Exception as e:
            messages.error(request, f"Error de conexión con la API: {e}")

    return redirect('gestion_pedidos')

def editar_producto(request):
    if request.method == "POST":
        producto_id = request.POST.get("producto_id")
        nombre = request.POST.get("nombre")
        descripcion = request.POST.get("descripcion")
        precio = request.POST.get("precio")
        stock = request.POST.get("stock")
        categoria_id = request.POST.get("categoria_id")
        marca_id = request.POST.get("marca_id")

        # Obtener la imagen actual desde la API (no se permite cambiarla)
        imagen = ""
        try:
            r = requests.get(f"http://localhost:3002/productos/{producto_id}")
            if r.status_code == 200:
                imagen = r.json().get("imagen", "")
        except Exception as e:
            messages.error(request, f"Error obteniendo imagen actual: {e}")

        payload = {
            "nombre": nombre,
            "descripcion": descripcion,
            "precio": float(precio),
            "stock": int(stock),
            "categoria_id": int(categoria_id),
            "marca_id": int(marca_id),
            "imagen": imagen  # Mantiene la imagen actual
        }

        try:
            r = requests.put(f"http://localhost:3002/put_productos/{producto_id}", json=payload)
            if r.status_code == 200:
                messages.success(request, "Producto actualizado correctamente.")
            else:
                messages.error(request, f"Error al actualizar producto: {r.json().get('error', 'Error desconocido')}")
        except Exception as e:
            messages.error(request, f"Error de conexión con la API: {e}")

    return redirect('gestion_pedidos')

def eliminar_producto(request):
    if request.method == "POST":
        producto_id = request.POST.get("producto_id")
        try:
            r = requests.delete(f"http://localhost:3002/del_productos/{producto_id}")
            if r.status_code == 200:
                messages.success(request, "Producto eliminado correctamente.")
            else:
                messages.error(request, f"Error al eliminar producto: {r.json().get('error', 'Error desconocido')}")
        except Exception as e:
            messages.error(request, f"Error de conexión con la API: {e}")
    return redirect('gestion_pedidos')

def registro_cliente(request):
    return render(request, "core/registro_cliente.html")

def historial_pedidos(request):
    usuario_id = request.session.get('usuario_id')
    pedidos = []
    if usuario_id:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    ped.pedido_id,
                    ped.fecha_pedido,
                    ped.total,
                    ep.nombre AS estado_pedido,
                    COALESCE(epa.nombre, 'Sin pago') AS estado_pago
                FROM PEDIDO ped
                JOIN ESTADO ep ON ped.estado_id = ep.estado_id
                LEFT JOIN PAGO p ON p.pedido_id = ped.pedido_id
                LEFT JOIN ESTADO_PAGO epa ON p.estado_pago_id = epa.estado_pago_id
                WHERE ped.cliente_id = %s
                ORDER BY ped.fecha_pedido DESC
            """, [usuario_id])
            pedidos = [
                {
                    "pedido_id": row[0],
                    "fecha_pedido": row[1],
                    "total": row[2],
                    "estado_pedido": row[3],
                    "estado_pago": row[4],
                }
                for row in cursor.fetchall()
            ]
    return render(request, "core/historial_pedidos.html", {"pedidos": pedidos})

def cambiar_password_admin(request):
    usuario_id = request.session.get('usuario_id')
    if not usuario_id:
        return redirect('login')
    usuario = Usuario.objects.get(id_usuario=usuario_id)
    if usuario.rol_id != 3:
        return redirect('home')

    if request.method == "POST":
        nueva = request.POST.get("nueva_password")
        confirmar = request.POST.get("confirmar_password")
        if not nueva or not confirmar:
            messages.error(request, "Debes completar ambos campos.")
        elif nueva != confirmar:
            messages.error(request, "Las contraseñas no coinciden.")
        else:
            usuario.password = nueva  # Usa hash si tienes
            usuario.debe_cambiar_password = 0
            usuario.save()
            messages.success(request, "Contraseña cambiada correctamente. ¡Bienvenido!")
            return redirect('admin_usuarios')
    return render(request, "core/cambiar_password_admin.html")

from django.db.models import Q
from .models import Producto, Categoria

def busqueda_productos(request):
    query = request.GET.get('q', '')
    categoria_id = request.GET.get('categoria')
    productos = Producto.objects.all()
    if query:
        productos = productos.filter(
            Q(nombre__icontains=query) | Q(descripcion__icontains=query)
        )
    if categoria_id:
        productos = productos.filter(categoria_id=categoria_id)
    categorias = Categoria.objects.all()
    return render(request, 'core/busqueda_productos.html', {
        'productos': productos,
        'categorias': categorias,
        'query': query,
        'categoria_seleccionada': int(categoria_id) if categoria_id else None,
    })

def detalle_producto(request, producto_id):
    from .models import Producto
    producto = get_object_or_404(Producto, producto_id=producto_id)
    todos_productos = Producto.objects.exclude(producto_id=producto_id)
    return render(request, 'core/detalle_producto.html', {
        'producto': producto,
        'todos_productos': todos_productos
    })